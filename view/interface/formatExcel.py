# coding:utf-8
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Optional, override

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtWidgets import QFileDialog, QHBoxLayout, QVBoxLayout, QWidget
from qfluentwidgets import BodyLabel, InfoBar, InfoBarPosition, PushButton, TextEdit

from common.config import cfg
from view.components.dropable_lineEdit import DropableLineEditDir
from view.interface.gallery_interface import GalleryInterface


class FormatWorker(QThread):
    logInfo = Signal(str)

    # 删除的列名
    # COLUMNS_TO_DROP = ["营业执照图片", "药品图片", "原价"]
    COLUMNS_TO_DROP = ["营业执照图片", "原价"]

    def __init__(self, excel_dir: Path):
        super().__init__()
        self.excel_dir = excel_dir

    def rename_cell(self, excel_path: Path):
        """重命名药品名称列"""
        df = pd.read_excel(excel_path)

        # 重命名药品名称列
        df["药品名称"] = "乐药师" + excel_path.stem

        df.to_excel(excel_path, index=False)

        self.logInfo.emit(f"{excel_path.stem} 重命名完成")

    def delete_col(self, excel_path: Path):
        """删除指定列"""
        df = pd.read_excel(excel_path)

        # 删除存在的列
        existing_columns = [col for col in self.COLUMNS_TO_DROP if col in df.columns]

        if not existing_columns:
            return

        df.drop(columns=existing_columns, inplace=True)

        df.to_excel(excel_path, index=False)

        self.logInfo.emit(f"{excel_path.stem} 删除列完成")

    def format_cell(self, excel_path: Path):
        """设置单元格格式"""
        wb = load_workbook(excel_path)
        ws = wb.active

        # 设置缩放为 100%
        ws.sheet_view.zoomScale = 100

        # 设置第一行字体及对齐方式
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=15, bold=True)

        # 设置第一行的行高
        ws.row_dimensions[1].height = 25

        # 定义列宽和对齐方式
        column_formats = {
            "A": 30,  # uuid
            "B": 45,  # 药店名称
            "C": 20,  # 店铺主页
            "D": 50,  # 资质名称
            "E": 35,  # 药品名称
            "F": 20,  # 挂网价格
            "G": 15,  # 平台
            "H": 14,  # 排查日期
        }

        for col, width in column_formats.items():
            ws.column_dimensions[col].width = width
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(excel_path)

        self.logInfo.emit(f"{excel_path.name} 格式化完成")

    def merge_excels(self):
        all_data: list[pd.DataFrame] = []
        for excel_file in self.excel_dir.glob("*.xlsx"):
            try:
                if any(keyword in excel_file.stem for keyword in ["~", "对照", "排查"]):
                    continue
                df = pd.read_excel(excel_file, dtype=str)
                all_data.append(df)
            except Exception as e:
                self.logInfo.emit(f"{excel_file.name} 读取失败: {e}")
                continue

            if not all_data:
                return

            merged_df = pd.concat(all_data, ignore_index=True)

            # 去除重复行
            merged_df.drop_duplicates(subset=["uuid"], inplace=True)

            # 保存到新文件
            merged_df.to_excel(self.excel_dir / "合并.xlsx", index=False)

    @override
    def run(self):
        """按顺序处理 Excel 文件"""
        futures = {}

        # 合并所有 Excel 文件
        self.merge_excels()

        with ThreadPoolExecutor(max_workers=8) as t:
            for excel_file in self.excel_dir.glob("*.xlsx"):
                try:
                    if any(
                        keyword in excel_file.stem for keyword in ["~", "对照", "排查"]
                    ):
                        continue

                    f = t.submit(self.delete_col, excel_file)
                    futures[f] = excel_file

                except Exception as e:
                    self.logInfo.emit(f"{self.excel_dir.name} 处理失败: {e}")
                    continue

            for f in futures:
                excel_file = futures[f]
                try:
                    f.result()
                    t.submit(self.format_cell, excel_file)
                except Exception as e:
                    self.logInfo.emit(f"{excel_file.name} 处理失败: {e}")
                    continue


class FormatExcelInterface(GalleryInterface):
    def __init__(self, parent=None):
        super().__init__(title="修改 Excel 文件的格式", parent=parent)

        self.view = QWidget(self)

        # 状态提示
        self.stateTooltip = None

        # 界面的垂直布局
        self.vBoxLayout = QVBoxLayout(self.view)
        self.hBoxLayout = QHBoxLayout()

        self.label_excel_path = BodyLabel(text="Excel 文件所在文件夹: ")

        # 显示 Excel 文件所在文件夹的文本框
        self.lineEdit_excel_path = DropableLineEditDir()
        self.lineEdit_excel_path.textChanged.connect(
            lambda: cfg.set(cfg.formatExcel_excel_path, self.lineEdit_excel_path.text())
        )
        self.lineEdit_excel_path.setPlaceholderText(
            "请选择或者拖入 Excel 文件所在文件夹"
        )

        # 选择路径的按钮
        self.btn_select_path = PushButton(text="···")
        self.btn_select_path.clicked.connect(
            lambda: self.lineEdit_excel_path.setText(
                QFileDialog.getExistingDirectory(self, "选择文件夹")
            )
        )

        # 刷新按钮
        self.btn_refresh = PushButton(text="开始")
        self.btn_refresh.clicked.connect(self.format)

        # 文本框 用于打印日志
        self.textEdit_log = TextEdit()
        self.textEdit_log.setPlaceholderText("此处是用来打印日志的")

        self.hBoxLayout.addWidget(self.label_excel_path)
        self.hBoxLayout.addWidget(self.lineEdit_excel_path)
        self.hBoxLayout.addWidget(self.btn_select_path)

        self.vBoxLayout.addLayout(self.hBoxLayout)

        self.vBoxLayout.addWidget(self.btn_refresh)
        self.vBoxLayout.addWidget(self.textEdit_log)

        self.__initWidget()

        self.worker: Optional[FormatWorker] = None

        self.lineEdit_excel_path.setText(cfg.formatExcel_excel_path.value)

    def __initWidget(self):
        self.view.setObjectName("")
        self.setObjectName("FormatExcelInterface")

        self.setWidget(self.view)
        self.setWidgetResizable(True)

    def __initLayout(self):
        self.hBoxLayout.setSpacing(8)

    def createErrorInfoBar(self, title, content):
        """
        创建错误信息栏
        :param title:
        :param content:
        :return:
        """
        InfoBar.error(
            title=title,
            content=content,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM_RIGHT,
            duration=2000,
            parent=self,
        )

    def createSuccessInfoBar(self, title, content):
        """
        创建成功信息栏
        :param title:
        :param content:
        :return:
        """
        InfoBar.success(
            title=title,
            content=content,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=2000,
            parent=self,
        )

    @Slot(str)
    def logInfo(self, info):
        """
        打印日志
        """
        self.textEdit_log.append(info)

    @Slot()
    def finished(self):
        """处理完成"""
        self.lineEdit_excel_path.setEnabled(True)
        self.btn_select_path.setEnabled(True)
        self.btn_refresh.setEnabled(True)

        self.createSuccessInfoBar("成功", "处理完成")

    def format(self):
        """格式化 Excel 文件"""
        self.textEdit_log.clear()

        excel_dir = Path(self.lineEdit_excel_path.text())

        if excel_dir.is_file():
            return

        if not excel_dir.exists():
            self.createErrorInfoBar("错误", "文件夹不存在")
            return

        self.lineEdit_excel_path.setEnabled(False)
        self.btn_select_path.setEnabled(False)
        self.btn_refresh.setEnabled(False)

        self.textEdit_log.clear()

        self.worker = FormatWorker(excel_dir)
        self.worker.logInfo.connect(self.logInfo)
        self.worker.finished.connect(self.finished)
        self.worker.start()
