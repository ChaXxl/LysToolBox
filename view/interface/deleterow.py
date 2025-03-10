# coding:utf-8
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import override

from openpyxl import load_workbook
from PySide6.QtCore import QThread, Signal, Slot
from PySide6.QtWidgets import QFileDialog, QHBoxLayout, QVBoxLayout, QWidget
from qfluentwidgets import BodyLabel, PushButton, TextEdit

from common.config import cfg
from view.components.dropable_lineEdit import DropableLineEditDir
from view.interface.gallery_interface import GalleryInterface


class DeleteRowWorker(QThread):
    logInfo = Signal(str)

    def __init__(self, root_dir: Path):
        super().__init__()

        self.root_dir = root_dir

        self.total_rows = 0
        self.deleted_rows = 0

    def getPicList(self, folder: Path) -> set:
        """获取文件夹下的图片列表"""
        try:
            pic_list = {img.stem.split("_")[-1] for img in folder.glob("*")}
            return pic_list
        except Exception as e:
            self.logInfo.emit(f"{folder.stem} 获取图片列表失败: {e}")
            return set()

    def process_excel(self, excel_path: Path, pic_list: set):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            # 收集需要删除的行
            rows_to_delete = []
            for row in ws.iter_rows(min_row=2):
                self.total_rows += 1

                uuid = row[0].value  # uuid

                if uuid and uuid not in pic_list and row[6].value:
                    self.deleted_rows += 1
                    rows_to_delete.append(row[0].row)

            if not rows_to_delete:
                self.logInfo.emit(f"{excel_path.stem} 无需删除")
                return

            # 删除行
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
                self.logInfo.emit(f"删除: {excel_path.stem} {row_idx}")

            # 保存文件
            wb.save(excel_path)
            self.logInfo.emit(f"{excel_path.stem} 处理完成")

        except Exception as e:
            self.logInfo.emit(f"{excel_path.stem} 处理失败: {e}")

    @override
    def run(self):
        start_time = datetime.now()

        with ThreadPoolExecutor() as t:
            futures = []

            # 只遍历文件夹
            for folder in self.root_dir.iterdir():
                if not folder.is_dir():
                    continue

                future = t.submit(self.getPicList, folder)
                futures.append((future, folder))

            # 等待图片列表任务完成, 并提交 Excel 处理任务
            for future, folder in futures:
                pic_list = future.result()
                if not pic_list:
                    continue

                excel_path = self.root_dir / f"{folder.stem}.xlsx"
                if not excel_path.exists():
                    self.logInfo.emit(f"{excel_path.stem} 不存在")
                    continue

                # 提交 Excel 处理任务
                t.submit(self.process_excel, excel_path, pic_list)

        self.logInfo.emit(
            f"\n共有 {self.total_rows} 行, 删除了 {self.deleted_rows} 行, 还剩 {self.total_rows - self.deleted_rows} 行"
        )

        self.logInfo.emit(f"\n耗时: {datetime.now() - start_time}")


class DeleteRowInterface(GalleryInterface):
    def __init__(self, parent=None):
        super().__init__(title="删除行", parent=parent)

        self.view = QWidget(self)

        # 状态提示
        self.stateTooltip = None

        self.vBoxLayout = QVBoxLayout(self.view)
        self.hBoxLayout = QHBoxLayout()
        self.hLayout_count = QHBoxLayout()

        self.label_excel_path = BodyLabel(text="Excel 文件所在文件夹: ")

        # 显示 Excel 文件所在文件夹的文本框
        self.lineEdit_excel_path = DropableLineEditDir()
        self.lineEdit_excel_path.setPlaceholderText(
            "请选择或者拖入 Excel 文件所在文件夹"
        )
        self.lineEdit_excel_path.textChanged.connect(
            lambda: cfg.set(cfg.deleteRow_excel_path, self.lineEdit_excel_path.text())
        )

        # 选择路径的按钮
        self.btn_select_path = PushButton(text="···")
        self.btn_select_path.clicked.connect(
            lambda: self.lineEdit_excel_path.setText(
                QFileDialog.getExistingDirectory(self, "选择文件夹")
            )
        )

        # 刷新按钮
        self.btn_refresh = PushButton(text="删除")
        self.btn_refresh.clicked.connect(self.delete)

        # 文本框 用于打印日志
        self.textEdit_log = TextEdit()
        self.textEdit_log.setPlaceholderText("此处是用来打印日志的")

        # 横向布局添加控件
        self.hBoxLayout.addWidget(self.label_excel_path)
        self.hBoxLayout.addWidget(self.lineEdit_excel_path)
        self.hBoxLayout.addWidget(self.btn_select_path)

        # 纵向布局添加布局
        self.vBoxLayout.addLayout(self.hBoxLayout)
        self.vBoxLayout.addWidget(self.btn_refresh)
        self.vBoxLayout.addLayout(self.hLayout_count)

        self.vBoxLayout.addWidget(self.textEdit_log)

        self.__initWidget()

        self.worker = None

        self.lineEdit_excel_path.setText(cfg.deleteRow_excel_path.value)

    def __initWidget(self):
        self.view.setObjectName("")
        self.setObjectName("DeleteRowInterface")

        self.setWidget(self.view)
        self.setWidgetResizable(True)

    @Slot()
    def logInfo(self, msg: str):
        """
        打印日志
        """
        self.textEdit_log.append(msg)

    def delete(self):
        """
        删除
        """
        self.textEdit_log.clear()

        root_dir = Path(self.lineEdit_excel_path.text())

        if not root_dir or not root_dir.exists():
            return

        self.worker = DeleteRowWorker(root_dir)
        self.worker.logInfo.connect(self.logInfo)

        self.worker.start()
