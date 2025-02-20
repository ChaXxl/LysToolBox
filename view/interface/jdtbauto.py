# coding:utf-8
from datetime import datetime
from pathlib import Path
from typing import Optional, override

from openpyxl.reader.excel import load_workbook
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtWidgets import QFileDialog, QHBoxLayout, QVBoxLayout, QWidget
from qfluentwidgets import (
    BodyLabel,
    InfoBar,
    InfoBarPosition,
    ProgressBar,
    PushButton,
    TextEdit,
)

from common.config import cfg
from utils.jd import JD
from utils.tb import TB
from view.components.dropable_lineEdit import DropableLineEditDir, DropableLineEditExcel
from view.interface.gallery_interface import GalleryInterface


class JdTbWorker(QThread):
    logInfo = Signal(str)
    setProgress = Signal(int)
    setProgressInfo = Signal(int, int)

    def __init__(self, keywords_path: Path, output_dir: Path):
        super().__init__()

        self.keywords_path = keywords_path
        self.output_dir = output_dir

        self.jd = JD(self.output_dir)
        self.tb = TB(self.output_dir)

        # 连接 JD 和 TB 中的 logInfo 信号到 JdTbWorker 的 logInfo 信号
        self.jd.logInfo = self.tb.logInfo = self.logInfo

    @override
    def run(self):
        start = datetime.now()

        if not self.keywords_path.exists():
            return

        wb = load_workbook(self.keywords_path, read_only=True, data_only=True)
        ws = wb.active

        keywords: [str] = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]

        for idx, keyword in enumerate(keywords):
            self.setProgress.emit((idx + 1) // len(keywords) * 100)
            self.setProgressInfo.emit(idx + 1, len(keywords))

            if Path(self.output_dir, f"{keyword}.xlsx").exists():
                self.logInfo.emit(f"{keyword} 已经存在，跳过")
                continue

            self.jd.search(keyword)

            self.tb.search(keyword)

        self.logInfo.emit(f"\n耗时: {datetime.now() - start}")


class JdTBbAutoInterface(GalleryInterface):
    def __init__(self, parent=None):
        super().__init__("京东淘宝自动化", parent=parent)

        self.view = QWidget(self)

        # 状态提示
        self.stateTooltip = None

        self.vBoxLayout = QVBoxLayout(self.view)
        self.hBoxLayout = QHBoxLayout()
        self.hBoxLayout_output = QHBoxLayout()
        self.hBoxLayout_progress = QHBoxLayout()

        self.label_onnx = BodyLabel(text="药品的 Excel 文件: ")
        # Excel 文件所在文件夹的文本框
        self.lineEdit_keywordPath = DropableLineEditExcel()
        self.lineEdit_keywordPath.setPlaceholderText("请选择或者拖入药品的 Excel 文件")
        self.lineEdit_keywordPath.textChanged.connect(
            lambda: cfg.set(cfg.jdtb_keyword_path, self.lineEdit_keywordPath.text())
        )

        self.label_output = BodyLabel(text="输出文件夹: ")
        # 输出文件夹的文本框
        self.lineEdit_output_path = DropableLineEditDir()
        self.lineEdit_output_path.setPlaceholderText("请选择或者拖入输出文件夹")
        self.lineEdit_output_path.textChanged.connect(
            lambda: cfg.set(cfg.jdtb_output_path, self.lineEdit_output_path.text())
        )

        # 选择药品 Excel 文件的按钮
        self.btn_select_keyword_path = PushButton(text="···")
        self.btn_select_keyword_path.clicked.connect(
            lambda: self.lineEdit_keywordPath.setText(
                QFileDialog.getOpenFileName(
                    self, "选择文件", filter="Excel 文件(*.xlsx)"
                )[0]
            )
        )

        # 选择输出路径的按钮
        self.btn_select_output_path = PushButton(text="···")
        self.btn_select_output_path.clicked.connect(
            lambda: self.lineEdit_output_path.setText(
                QFileDialog.getExistingDirectory(self, "选择文件夹")
            )
        )

        # 下载按钮
        self.btn_download = PushButton(text="开始")
        self.btn_download.clicked.connect(self.start)

        # 文本框 用于打印日志
        self.textEdit_log = TextEdit()
        self.textEdit_log.setPlaceholderText("此处是用来打印日志的")

        # 进度条
        self.progressBar = ProgressBar()

        # 进度提示标签
        self.label_progress = BodyLabel(text="0/0")

        self.hBoxLayout.addWidget(self.label_onnx)
        self.hBoxLayout.addWidget(self.lineEdit_keywordPath)
        self.hBoxLayout.addWidget(self.btn_select_keyword_path)

        self.hBoxLayout_output.addWidget(self.label_output)
        self.hBoxLayout_output.addWidget(self.lineEdit_output_path)
        self.hBoxLayout_output.addWidget(self.btn_select_output_path)

        self.hBoxLayout_progress.addWidget(self.progressBar)
        self.hBoxLayout_progress.addWidget(self.label_progress)

        self.vBoxLayout.addLayout(self.hBoxLayout)
        self.vBoxLayout.addLayout(self.hBoxLayout_output)

        self.vBoxLayout.addWidget(self.btn_download)
        self.vBoxLayout.addWidget(self.textEdit_log)

        # 进度条布局
        self.vBoxLayout.addLayout(self.hBoxLayout_progress)

        self.__initWidget()

        # 从配置文件中读取路径
        self.lineEdit_keywordPath.setText(cfg.jdtb_keyword_path.value)
        self.lineEdit_output_path.setText(cfg.jdtb_output_path.value)

        self.worker: Optional[JdTbWorker] = None

    def __initWidget(self):
        self.view.setObjectName("")
        self.setObjectName("JdTBbAutoInterface")

        self.setWidget(self.view)
        self.setWidgetResizable(True)

    def createErrorInfoBar(self, title, content):
        """
        创建错误信息栏
        """
        InfoBar.error(
            title=title,
            content=content,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM_RIGHT,
            duration=2000,
            parent=self,
        )

    def createSuccessInfoBar(self, title, content):
        """
        创建错误信息栏
        """
        InfoBar.success(
            title=title,
            content=content,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=2000,
            parent=self,
        )

    @Slot(str)
    def logInfo(self, info):
        """
        打印日志
        """
        self.textEdit_log.append(info)

    @Slot(int)
    def setProgress(self, value):
        """
        设置进度条进度
        """
        self.progressBar.setValue(value)

    @Slot(int, int)
    def setProgressInfo(self, value, total):
        """
        设置进度条进度
        """
        self.label_progress.setText(f"{value}/{total}")

    @Slot()
    def finished(self):
        self.lineEdit_keywordPath.setEnabled(True)
        self.btn_select_keyword_path.setEnabled(True)

        self.lineEdit_output_path.setEnabled(True)
        self.btn_select_output_path.setEnabled(True)

        self.btn_download.setEnabled(True)

        if self.stateTooltip is not None:
            self.stateTooltip.hide()

        self.createSuccessInfoBar("完成", "完成 ✅")

    def start(self):
        self.textEdit_log.clear()

        # 检查是否选择了 药品 Excel 文件
        keywords_path = self.btn_select_keyword_path.text()
        if not keywords_path:
            self.createErrorInfoBar("错误", "请选择 onnx 模型文件")
            return

        # 检查是否选择了输出文件夹
        output_dir = self.lineEdit_output_path.text()
        if not output_dir:
            self.createErrorInfoBar("错误", "请选择输出文件夹")
            return

        keywords_path = Path(self.lineEdit_keywordPath.text())
        output_dir = Path(self.lineEdit_output_path.text())

        self.lineEdit_keywordPath.setEnabled(False)
        self.btn_select_keyword_path.setEnabled(False)

        self.lineEdit_output_path.setEnabled(False)
        self.btn_select_output_path.setEnabled(False)

        self.btn_download.setEnabled(False)

        self.worker = JdTbWorker(keywords_path, output_dir)

        self.worker.logInfo.connect(self.logInfo)
        self.worker.finished.connect(self.finished)
        self.worker.setProgress.connect(self.setProgress)
        self.worker.setProgressInfo.connect(self.setProgressInfo)

        self.worker.start()
