from pathlib import Path
from typing import List, Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import polars as pl
import shortuuid
from loguru import logger
from PySide6.QtCore import Signal


class Save:
    logInfo = Signal(str)

    def __init__(self): ...

    def to_excel(
        self, filename: Path, datas: List[List[Any]], tag: Optional[str] = None
    ) -> None:
        """
        保存数据到Excel文件

        Args:
            filename: 保存路径
            datas: 要保存的数据列表
            tag: 标签-指明哪个平台
        """
        # 数据为空则直接返回
        if not datas:
            return

        headers = [
            "uuid",
            "药店名称",
            "店铺主页",
            "资质名称",
            "药品名",
            "药品ID",
            "药品图片",
            "挂网价格",
            "平台",
            "排查日期",
        ]

        new_data = pl.DataFrame(datas, schema=headers, orient="row")
        existing_df: Optional[pl.DataFrame] = None

        #  如果文件存在, 读取数据并去重
        if filename.exists():
            try:
                existing_df = pl.read_excel(filename)
            except Exception as e:
                logger.error(f"读取Excel文件失败: {e}")
                self.logInfo.emit(f"读取Excel文件失败: {e}\n请检查文件格式或路径")
                return

            # 对齐数据类型, 全部转换为字符串
            existing_df = existing_df.with_columns(pl.all().cast(pl.Utf8))
            new_data = new_data.with_columns(pl.all().cast(pl.Utf8))

            # 去重
            combined_df = pl.concat([existing_df, new_data], how="vertical")
            combined_df = combined_df.unique(
                subset=["药店名称", "店铺主页", "药品名", "挂网价格", "平台"]
            )

        else:
            combined_df = new_data

        # 按平台列升序排序
        combined_df = combined_df.sort("平台")

        # 保存数据到Excel
        try:
            combined_df.write_excel(filename)
        except Exception as e:
            logger.error(f"保存数据到Excel失败: {e}")
            self.logInfo.emit(f"保存数据到Excel失败: {e}\n请检查文件格式或路径")
            return

        saved_count = (
            combined_df.shape[0] - existing_df.shape[0]
            if existing_df is not None
            else combined_df.shape[0]
        )

        msg = f"\n\n{filename.stem} {tag}-保存了 {saved_count} 条, 数据总条数: {combined_df.shape[0]}\n\n"
        self.logInfo.emit(msg)

        # 格式化 Excel 文件
        try:
            wb = load_workbook(filename)
            ws = wb.active

            # 设置缩放为 100%
            ws.sheet_view.zoomScale = 100

            # 设置第一行字体及对齐方式
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=15, bold=True)

            # 设置第一行的行高
            ws.row_dimensions[1].height = 25

            # 定义列宽和对齐方式
            column_formats = {
                "A": 30,  # uuid
                "B": 45,  # 药店名称
                "C": 20,  # 店铺主页
                "D": 50,  # 资质名称
                "E": 35,  # 药品名称
                "F": 15,  # 药品ID
                "G": 20,  # 药品图片
                "H": 23,  # 挂网价格
                "I": 15,  # 平台
                "J": 18,  # 排查日期
            }

            for col, width in column_formats.items():
                ws.column_dimensions[col].width = width
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

            # 保存格式化后的文件
            wb.save(filename)
        except Exception as e:
            self.logInfo.emit(f"格式化Excel文件失败: {e}")
            logger.error(f"格式化Excel文件失败: {e}")
